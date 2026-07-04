#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
TRANSFORMADOR EXCEL -> JSON
Listado Maestro 2026 -> JSON limpio para el Importador PMO

- NO modifica el Excel de origen (solo lectura).
- NO toca backend / Prisma / base de datos / frontend.
- Analiza automáticamente todas las hojas del archivo.
- El Tipo Documental se deriva del NOMBRE DE LA HOJA (nunca de una columna).
- Genera:
    1) documentos-biblioteca.json  -> registros limpios
    2) errores-importacion.json    -> inconsistencias detectadas (no detiene el proceso)

Uso:
    python3 transformador.py <ruta_excel> [carpeta_salida]

Si no se indican argumentos, usa los valores por defecto definidos abajo.
"""

import sys
import os
import json
import re
import unicodedata
import datetime
from openpyxl import load_workbook

# =====================================================================
# CONFIGURACION POR DEFECTO
# =====================================================================

from pathlib import Path

BASE_DIR = Path(__file__).parent

DEFAULT_INPUT = BASE_DIR / "inputs" / "Listado_Maestro_2026__3_.xlsx"
DEFAULT_OUTPUT_DIR = BASE_DIR

VERSION_TRANSFORMADOR = "transformador-1.0.0"

# Fila donde estan los encabezados (fila 1 = titulos/instrucciones, fila 2 = encabezados reales)
HEADER_ROW = 2
FIRST_DATA_ROW = 3

# Layout de columnas COMUN a todas las hojas (posicion -> campo interno).
# Se usa POSICION en vez de nombre de encabezado porque algunas hojas
# tienen encabezados corruptos/con errores de tipeo (ej. hoja de
# Procedimientos), pero el ORDEN de columnas es idéntico en todas las hojas.
COLUMN_LAYOUT = {
    1:  "codigoDocumento",              # Codigo Manual / Codigo Documento
    2:  "codigoArea",                   # Codigo Area
    3:  "area",                         # Area / Cargo
    4:  "nombre",                       # Nombre del Documento
    5:  "proceso",                      # Proceso Asociado
    6:  "version",                      # Version
    7:  "fechaCreacion",                # Fecha de Creacion
    8:  "fechaRevision",                # Fecha ultima actualizacion
    9:  "fechaProximaRevision",         # Fecha proxima revision
    10: "diasRestantes",                # Dias restantes (informativo, no va en la salida)
    11: "vigencia",                     # Vigencia
    12: "estado",                       # Estado del Documento
    13: "observaciones",                # Observaciones
    14: "encargado1",                   # Encargado 1 (informativo, no va en la salida)
    15: "voBo",                         # VoBo (informativo, no va en la salida)
    16: "enlaceDrive",                  # Enlace Doc al drive (informativo, no va en la salida)
    17: "cargoActual",                  # Cargo Actual (informativo, no va en la salida)
    18: "responsableActualizacion",     # Responsable Actualizacion
    19: "responsableRevision",          # Responsable revision/aprobacion
    20: "tipoDocumentoExcel",           # Tipo de Documento (columna propia del excel, informativo)
}

# Encabezados esperados por posicion (para detectar "hoja con columnas inesperadas").
# Se comparan de forma flexible (sin tildes, sin mayusculas, ignorando espacios extra).
EXPECTED_HEADERS = {
    1:  "codigo manual",
    2:  "codigo area",
    3:  "area",              # aparece como "Area / Cargo" o "Area"
    4:  "nombre del documento",
    5:  "proceso asociado",
    6:  "version",
    7:  "fecha de creacion",
    8:  "fecha ultima actualizacion",
    9:  "fecha proxima revision",
    11: "vigencia",
    12: "estado del documento",
    13: "observaciones",
    18: "responsable actualizacion",
    20: "tipo de documento",
}

# Valores que jamas deben escribirse como responsable "inventado".
# Si el Excel trae literalmente alguno de estos textos como responsable,
# se guarda null y se registra como inconsistencia.
PLACEHOLDER_VALUES = {
    "sin asignar", "pendiente", "n/a", "na", "no aplica",
    "por definir", "por asignar", "sin definir", "-", "--",
}

# =====================================================================
# REGLAS: NOMBRE DE HOJA -> TIPO DOCUMENTAL
# =====================================================================
# Orden de evaluacion: la primera coincidencia gana.
# Se busca por palabra clave dentro del nombre de hoja ya normalizado
# (sin tildes, en minusculas). Esto permite detectar automaticamente
# tipos documentales nuevos si el nombre de la hoja los sugiere.

TIPO_DOCUMENTAL_RULES = [
    ("procedimiento", "PROCEDIMIENTO"),
    ("manual",        "MANUAL"),
    ("proceso",       "PROCESO"),
    ("instructivo",   "INSTRUCTIVO"),
    ("politica",      "POLITICA"),
    ("formato",       "FORMATO"),
]

# Reglas adicionales por combinacion de palabras (cuando el nombre de hoja
# no contiene una sola palabra clave clara, ej. "FOR y MZ").
TIPO_DOCUMENTAL_COMBO_RULES = [
    (["for", "mz"], "FORMATO"),
]


def normalizar_texto(valor):
    """Minusculas, sin tildes, espacios colapsados. None-safe."""
    if valor is None:
        return ""
    texto = str(valor).strip().lower()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    texto = re.sub(r"\s+", " ", texto)
    return texto


def detectar_tipo_documental(nombre_hoja):
    """Determina el Tipo Documental a partir del nombre de la hoja (nunca de una columna)."""
    hoja_norm = normalizar_texto(nombre_hoja)

    for palabra_clave, tipo in TIPO_DOCUMENTAL_RULES:
        if palabra_clave in hoja_norm:
            return tipo

    for palabras, tipo in TIPO_DOCUMENTAL_COMBO_RULES:
        if all(p in hoja_norm for p in palabras):
            return tipo

    # Si no se pudo determinar automaticamente, se marca explicitamente
    # en vez de inventar un tipo (no se debe asumir nada).
    return "DESCONOCIDO"


# =====================================================================
# UTILIDADES DE LIMPIEZA
# =====================================================================

def limpiar_texto(valor):
    """Convierte a string limpio (trim). Vacio/None -> None."""
    if valor is None:
        return None
    if isinstance(valor, str):
        texto = valor.strip()
        return texto if texto != "" else None
    # numeros u otros tipos que vengan en columnas de texto
    return str(valor).strip()


def limpiar_codigo_area(valor):
    """
    Codigo Area puede venir como numero (0, 1, 2) o como texto ('0.1', '0.A').
    Se conserva el valor tal cual, normalizado a string, sin forzar formato.
    """
    if valor is None:
        return None
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return str(valor).strip()


def limpiar_version(valor):
    """La version se mantiene EXACTAMENTE como viene en el Excel."""
    if valor is None:
        return None
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return str(valor).strip() if isinstance(valor, str) else valor


def parsear_fecha(valor):
    """
    Devuelve (fecha_iso_o_None, es_invalida).
    - Vacio -> (None, False)
    - datetime/date valido -> (YYYY-MM-DD, False)
    - Texto no interpretable como fecha -> (None, True)  [inconsistencia]
    """
    if valor is None:
        return None, False
    if isinstance(valor, (datetime.datetime, datetime.date)):
        if isinstance(valor, datetime.datetime):
            return valor.date().isoformat(), False
        return valor.isoformat(), False
    if isinstance(valor, str):
        texto = valor.strip()
        if texto == "":
            return None, False
        # Intentar formatos comunes antes de declarar invalida
        formatos = ["%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"]
        for fmt in formatos:
            try:
                return datetime.datetime.strptime(texto, fmt).date().isoformat(), False
            except ValueError:
                continue
        return None, True
    # Cualquier otro tipo inesperado (ej. numero suelto que no es fecha excel)
    return None, True


def limpiar_responsable(valor):
    """
    No se inventan responsables. Vacio -> None.
    Si el valor es un placeholder tipo 'Sin asignar'/'Pendiente'/'N/A',
    se guarda None y se reporta como inconsistencia.
    Devuelve (valor_limpio_o_None, es_placeholder_detectado).
    """
    texto = limpiar_texto(valor)
    if texto is None:
        return None, False
    if normalizar_texto(texto) in PLACEHOLDER_VALUES:
        return None, True
    return texto, False


def fila_esta_vacia(fila_dict):
    """Una fila se considera vacia si no tiene codigo, nombre, area ni codigo de area."""
    campos_clave = ["codigoDocumento", "nombre", "area", "codigoArea"]
    return all(limpiar_texto(fila_dict.get(c)) is None for c in campos_clave)


# =====================================================================
# VALIDACION DE ENCABEZADOS POR HOJA
# =====================================================================

def validar_encabezados(ws, nombre_hoja, errores):
    fila_headers = [ws.cell(row=HEADER_ROW, column=c).value for c in range(1, 21)]
    for pos, esperado in EXPECTED_HEADERS.items():
        real = normalizar_texto(fila_headers[pos - 1])
        # comparacion flexible: el esperado debe estar contenido en el real
        # (para tolerar variantes como "Area / Cargo" vs "Area")
        if esperado not in real and real not in esperado:
            errores.append({
                "hoja": nombre_hoja,
                "fila": HEADER_ROW,
                "campo": f"columna_{pos}",
                "tipo": "Hoja con columnas inesperadas",
                "detalle": f"Encabezado esperado ~'{esperado}', encontrado '{fila_headers[pos - 1]}'",
            })


# =====================================================================
# TRANSFORMACION PRINCIPAL
# =====================================================================

def transformar(ruta_excel):
    wb = load_workbook(ruta_excel, data_only=True)

    documentos = []
    errores = []
    codigos_vistos = {}  # codigoDocumento -> primera ubicacion (hoja, fila)

    for nombre_hoja in wb.sheetnames:
        ws = wb[nombre_hoja]
        tipo_documental = detectar_tipo_documental(nombre_hoja)

        if tipo_documental == "DESCONOCIDO":
            errores.append({
                "hoja": nombre_hoja,
                "fila": None,
                "campo": "tipoDocumental",
                "tipo": "Tipo documental no identificado por nombre de hoja",
                "detalle": f"No se pudo derivar el tipo documental del nombre de hoja '{nombre_hoja}'",
            })

        validar_encabezados(ws, nombre_hoja, errores)

        max_row = ws.max_row
        for fila_num in range(FIRST_DATA_ROW, max_row + 1):
            crudos = {}
            for pos, campo in COLUMN_LAYOUT.items():
                crudos[campo] = ws.cell(row=fila_num, column=pos).value

            if fila_esta_vacia(crudos):
                continue  # fila realmente vacia: se omite sin generar error

            registro = {}

            # --- Codigo de Documento (unico) ---
            codigo_doc = limpiar_texto(crudos.get("codigoDocumento"))
            registro["codigoDocumento"] = codigo_doc
            if codigo_doc is None:
                errores.append({
                    "hoja": nombre_hoja, "fila": fila_num, "campo": "codigoDocumento",
                    "tipo": "Documento sin codigo", "detalle": None,
                })
            else:
                if codigo_doc in codigos_vistos:
                    hoja_prev, fila_prev = codigos_vistos[codigo_doc]
                    errores.append({
                        "hoja": nombre_hoja, "fila": fila_num, "campo": "codigoDocumento",
                        "tipo": "Codigo duplicado",
                        "detalle": f"'{codigo_doc}' ya usado en hoja '{hoja_prev}' fila {fila_prev}",
                    })
                else:
                    codigos_vistos[codigo_doc] = (nombre_hoja, fila_num)

            # --- Codigo de Area (puede repetirse, no se valida unicidad) ---
            registro["codigoArea"] = limpiar_codigo_area(crudos.get("codigoArea"))

            # --- Tipo Documental (derivado de la hoja) ---
            registro["tipoDocumental"] = tipo_documental

            # --- Nombre ---
            registro["nombre"] = limpiar_texto(crudos.get("nombre"))
            if registro["nombre"] is None:
                errores.append({
                    "hoja": nombre_hoja, "fila": fila_num, "campo": "nombre",
                    "tipo": "Nombre de documento vacio", "detalle": None,
                })

            # --- Proceso ---
            registro["proceso"] = limpiar_texto(crudos.get("proceso"))
            if registro["proceso"] is None:
                errores.append({
                    "hoja": nombre_hoja, "fila": fila_num, "campo": "proceso",
                    "tipo": "Proceso vacio", "detalle": None,
                })

            # --- Area ---
            registro["area"] = limpiar_texto(crudos.get("area"))
            if registro["area"] is None:
                errores.append({
                    "hoja": nombre_hoja, "fila": fila_num, "campo": "area",
                    "tipo": "Area vacia", "detalle": None,
                })

            # --- Version (se conserva exactamente) ---
            registro["version"] = limpiar_version(crudos.get("version"))

            # --- Estado (tal cual viene, sin inventar) ---
            registro["estado"] = limpiar_texto(crudos.get("estado"))

            # --- Vigencia (tal cual viene) ---
            registro["vigencia"] = limpiar_texto(crudos.get("vigencia"))

            # --- Fechas ---
            for campo_fecha in ("fechaCreacion", "fechaRevision", "fechaProximaRevision"):
                valor_fecha, invalida = parsear_fecha(crudos.get(campo_fecha))
                registro[campo_fecha] = valor_fecha
                if invalida:
                    errores.append({
                        "hoja": nombre_hoja, "fila": fila_num, "campo": campo_fecha,
                        "tipo": "Fecha invalida",
                        "detalle": f"Valor original: '{crudos.get(campo_fecha)}'",
                    })

            # --- Responsables (nunca inventados) ---
            resp_act, placeholder_act = limpiar_responsable(crudos.get("responsableActualizacion"))
            registro["responsableActualizacion"] = resp_act
            if placeholder_act:
                errores.append({
                    "hoja": nombre_hoja, "fila": fila_num, "campo": "responsableActualizacion",
                    "tipo": "Responsable inexistente",
                    "detalle": f"Valor placeholder original: '{crudos.get('responsableActualizacion')}'",
                })

            resp_rev, placeholder_rev = limpiar_responsable(crudos.get("responsableRevision"))
            registro["responsableRevision"] = resp_rev
            if placeholder_rev:
                errores.append({
                    "hoja": nombre_hoja, "fila": fila_num, "campo": "responsableRevision",
                    "tipo": "Responsable inexistente",
                    "detalle": f"Valor placeholder original: '{crudos.get('responsableRevision')}'",
                })

            # --- Observaciones (tal cual) ---
            registro["observaciones"] = limpiar_texto(crudos.get("observaciones"))

            # --- Origen ---
            registro["origenHoja"] = nombre_hoja

            documentos.append(registro)

    return documentos, errores


# =====================================================================
# MAIN
# =====================================================================

def main():
    ruta_excel = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_INPUT
    carpeta_salida = sys.argv[2] if len(sys.argv) > 2 else DEFAULT_OUTPUT_DIR

    if not os.path.isfile(ruta_excel):
        print(f"ERROR: no se encontro el archivo de entrada: {ruta_excel}")
        sys.exit(1)

    os.makedirs(carpeta_salida, exist_ok=True)

    documentos, errores = transformar(ruta_excel)

    ruta_documentos = os.path.join(carpeta_salida, "documentos-biblioteca.json")
    ruta_errores = os.path.join(carpeta_salida, "errores-importacion.json")

    with open(ruta_documentos, "w", encoding="utf-8") as f:
        json.dump(documentos, f, ensure_ascii=False, indent=2)

    with open(ruta_errores, "w", encoding="utf-8") as f:
        json.dump(errores, f, ensure_ascii=False, indent=2)

    # Resumen en consola
    print("=" * 70)
    print("TRANSFORMACION COMPLETADA")
    print("=" * 70)
    print(f"Archivo de origen : {ruta_excel}")
    print(f"Documentos generados: {len(documentos)}")
    print(f"Inconsistencias detectadas: {len(errores)}")

    por_tipo = {}
    for d in documentos:
        por_tipo[d["tipoDocumental"]] = por_tipo.get(d["tipoDocumental"], 0) + 1
    print("\nDocumentos por tipo documental:")
    for tipo, cantidad in sorted(por_tipo.items()):
        print(f"  {tipo}: {cantidad}")

    por_tipo_error = {}
    for e in errores:
        por_tipo_error[e["tipo"]] = por_tipo_error.get(e["tipo"], 0) + 1
    print("\nInconsistencias por tipo:")
    for tipo, cantidad in sorted(por_tipo_error.items(), key=lambda x: -x[1]):
        print(f"  {tipo}: {cantidad}")

    print(f"\nArchivos generados:")
    print(f"  {ruta_documentos}")
    print(f"  {ruta_errores}")


if __name__ == "__main__":
    main()
